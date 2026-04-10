"""Excel workbook export helpers for grading results."""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import settings

logger = logging.getLogger(__name__)

HEADERS = [
    "Student ID",
    "Name",
    "Grade",
    "Compilation",
    "Tests Passed",
    "Banned Functions",
    "Similarity Score",
    "Notes",
]


def workbook_path_for_assignment(assignment_name: str) -> Path:
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "-", assignment_name.strip()).strip("-") or "assignment"
    export_dir = Path(settings.GRADE_EXPORT_DIR).expanduser()
    export_dir.mkdir(parents=True, exist_ok=True)
    return (export_dir / f"{safe_name}.xlsx").resolve()


async def write_grades(assignment_name: str, results: list[dict]) -> Path:
    """Write grading results to a local xlsx workbook."""
    workbook_path = workbook_path_for_assignment(assignment_name)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = assignment_name[:31] or "Grades"
    worksheet.append(HEADERS)

    for result in results:
        worksheet.append([
            result.get("student_id", ""),
            result.get("name", ""),
            result.get("grade", ""),
            result.get("compilation", ""),
            result.get("tests_passed", ""),
            result.get("banned_functions", ""),
            result.get("similarity_score", ""),
            result.get("notes", ""),
        ])

    workbook.save(workbook_path)
    logger.info("Wrote %d grades to %s", len(results), workbook_path)
    return workbook_path


async def update_grade(
    assignment_name: str,
    student_id: str,
    new_grade: float | str,
    reason: str,
) -> Path | None:
    """Update a student's grade in the workbook and append a note."""
    workbook_path = workbook_path_for_assignment(assignment_name)
    if not workbook_path.exists():
        logger.warning("Workbook not found for assignment '%s'", assignment_name)
        return None

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2):
        if row[0].value != student_id:
            continue

        row[2].value = new_grade
        current_notes = str(row[7].value or "").strip()
        note = f"Grade bumped to {new_grade}: {reason}"
        row[7].value = f"{current_notes}; {note}".strip("; ")
        workbook.save(workbook_path)
        logger.info("Updated %s in %s", student_id, workbook_path)
        return workbook_path

    logger.warning("Student %s not found in %s", student_id, workbook_path)
    return workbook_path


async def get_student(assignment_name: str, student_id: str) -> dict | None:
    """Load one student's row from the workbook."""
    workbook_path = workbook_path_for_assignment(assignment_name)
    if not workbook_path.exists():
        return None

    workbook = load_workbook(workbook_path, read_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] != student_id:
            continue
        student = {
            header: value
            for header, value in zip(HEADERS, row, strict=False)
            if value not in (None, "")
        }
        workbook.close()
        return student

    workbook.close()
    return None
